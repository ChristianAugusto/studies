from openpyxl import load_workbook


wb = load_workbook('./mercado hortifruit.xlsx')

# len(wb.worksheets) # quantidade de planilhas no arquivo

# for index, sheet in enumerate(wb.worksheets):
#     print(f'Sheet Index:[{index}], Title: {sheet.title}')

# for i in range(0, len(wb.worksheets)):
#     print(f'Sheet Index: {i}, Title: {wb.worksheets[i].title}')

# for sheet in wb.worksheets:
#     print(sheet.title)


worksheet = wb.worksheets[0]

for row_cells in worksheet.iter_rows():
    for cell in row_cells:
        print(f'cell.value = {cell.value}')
